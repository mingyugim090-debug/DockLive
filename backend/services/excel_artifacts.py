"""Excel workspace artifacts for the desktop-first Inline AI workflow.

The planner only uses parsed workspace facts and uploaded table values. The
renderer turns that plan into an XLSX workbook deterministically with openpyxl.
"""

from __future__ import annotations

import os
import re
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from core.errors import AnalysisError
from models.schemas import (
    DocumentWorkspace,
    ProjectFile,
    SheetTable,
    WorkbookChartPlan,
    WorkbookPlan,
    WorkbookSheetPlan,
    WorkbookSyncState,
    WorkbookTablePlan,
    WorkspaceArtifact,
)
from services.block_transforms import parse_number

DASHBOARD = "\ub300\uc2dc\ubcf4\ub4dc"
DOCUMENTS = "\uc81c\ucd9c\uc11c\ub958"
CHARTS = "\ucc28\ud2b8"
EVIDENCE = "\uc6d0\ubb38\uadfc\uac70"

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ARTIFACT_ROOT = Path(__file__).resolve().parents[1] / ".livedock_storage" / "artifacts"


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _safe_filename(title: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|\r\n]+', "_", (title or "excel-dashboard").strip())
    return (cleaned[:80] or "excel-dashboard") + ".xlsx"


def _analysis_ref(workspace: DocumentWorkspace, field: str) -> str:
    analysis_id = workspace.analysis.id if workspace.analysis else "no-analysis"
    return f"analysis:{analysis_id}:{field}"


def _first_spreadsheet(files: list[ProjectFile]) -> tuple[ProjectFile, int, SheetTable] | None:
    for project_file in files:
        if not project_file.sheet_data:
            continue
        for index, sheet in enumerate(project_file.sheet_data.sheets):
            if sheet.headers:
                return project_file, index, sheet
    return None


def _sheet_source_ref(project_file: ProjectFile, index: int) -> str:
    return f"{project_file.id}:sheet-{index}"


def _numeric_chart_from_sheet(project_file: ProjectFile, index: int, sheet: SheetTable) -> WorkbookChartPlan | None:
    rows = sheet.rows
    if not sheet.headers or not rows:
        return None
    labels: list[str] = []
    values: list[float] = []
    for row in rows:
        label = row[0] if row else ""
        value_text = row[1] if len(row) > 1 else ""
        value = parse_number(value_text)
        if label and value is not None:
            labels.append(label)
            values.append(value)
    if not labels or not values:
        return None
    return WorkbookChartPlan(
        id=f"wc-{uuid4()}",
        title=f"{sheet.name or project_file.filename} chart",
        anchor="A3",
        labels=labels,
        values=values,
        series_name=sheet.headers[1] if len(sheet.headers) > 1 else "value",
        source_ref=_sheet_source_ref(project_file, index),
    )


def _analysis_dashboard_rows(workspace: DocumentWorkspace) -> list[list[str]]:
    analysis = workspace.analysis
    if analysis is None:
        return [["field", "value", "source"], ["title", "", "no_information"]]
    deadline = next((item for item in analysis.timeline if item.is_deadline), None)
    return [
        ["field", "value", "source"],
        ["title", analysis.title, _analysis_ref(workspace, "title")],
        ["organization", analysis.organization, _analysis_ref(workspace, "organization")],
        ["summary", analysis.summary, _analysis_ref(workspace, "summary")],
        ["deadline", deadline.date if deadline else "", _analysis_ref(workspace, "timeline")],
        ["applicant_kind", analysis.applicant_kind, _analysis_ref(workspace, "applicant_kind")],
    ]


def _documents_rows(workspace: DocumentWorkspace) -> list[list[str]]:
    rows = [["document", "required", "source"]]
    analysis = workspace.analysis
    if analysis is None or not analysis.checklist:
        rows.append(["", "confirmation_required", "no_information"])
        return rows
    for item in analysis.checklist:
        rows.append([item.label, item.category, _analysis_ref(workspace, f"checklist:{item.id}")])
    return rows


def _evidence_rows(workspace: DocumentWorkspace) -> list[list[str]]:
    rows = [["source", "type", "filename", "note"]]
    if workspace.analysis:
        rows.append([workspace.analysis.id, "analysis", workspace.analysis.title, "source-grounded analysis"])
    for project_file in workspace.files:
        rows.append([project_file.id, project_file.file_kind, project_file.filename, "; ".join(project_file.warnings)])
    return rows


def build_workbook_plan(workspace: DocumentWorkspace) -> WorkbookPlan:
    spreadsheet = _first_spreadsheet(workspace.files)
    chart = _numeric_chart_from_sheet(*spreadsheet) if spreadsheet else None
    confirmation_required = []
    if workspace.analysis is None:
        confirmation_required.append("source confirmation: notice analysis is required before final use")
    if not spreadsheet:
        confirmation_required.append("source confirmation: upload a spreadsheet/table file for chart data")
    if not chart:
        confirmation_required.append("source confirmation: no numeric source table was available for chart generation")
    confirmation_required.append("source confirmation: review exported cells before final submission")

    source_table = None
    if spreadsheet:
        project_file, index, sheet = spreadsheet
        source_table = WorkbookTablePlan(
            id=f"wt-{uuid4()}",
            title=sheet.name or project_file.filename,
            anchor="A3",
            headers=list(sheet.headers),
            rows=[list(row) for row in sheet.rows],
            source_ref=_sheet_source_ref(project_file, index),
        )

    sheets = [
        WorkbookSheetPlan(
            id="dashboard",
            name=DASHBOARD,
            title=workspace.analysis.title if workspace.analysis else workspace.title or DASHBOARD,
            tables=[
                WorkbookTablePlan(
                    id=f"wt-{uuid4()}",
                    title="notice summary",
                    anchor="A3",
                    headers=["field", "value", "source"],
                    rows=_analysis_dashboard_rows(workspace)[1:],
                    source_ref=_analysis_ref(workspace, "summary"),
                )
            ],
        ),
        WorkbookSheetPlan(
            id="documents",
            name=DOCUMENTS,
            title=DOCUMENTS,
            tables=[
                WorkbookTablePlan(
                    id=f"wt-{uuid4()}",
                    title="submission documents",
                    anchor="A3",
                    headers=["document", "required", "source"],
                    rows=_documents_rows(workspace)[1:],
                    source_ref=_analysis_ref(workspace, "checklist"),
                )
            ],
        ),
        WorkbookSheetPlan(
            id="charts",
            name=CHARTS,
            title=CHARTS,
            tables=[source_table] if source_table else [],
            charts=[chart] if chart else [],
        ),
        WorkbookSheetPlan(
            id="evidence",
            name=EVIDENCE,
            title=EVIDENCE,
            tables=[
                WorkbookTablePlan(
                    id=f"wt-{uuid4()}",
                    title="source evidence",
                    anchor="A3",
                    headers=["source", "type", "filename", "note"],
                    rows=_evidence_rows(workspace)[1:],
                    source_ref=_analysis_ref(workspace, "source_evidence"),
                )
            ],
        ),
    ]
    return WorkbookPlan(
        id=f"wp-{uuid4()}",
        title=workspace.analysis.title if workspace.analysis else workspace.title or "Excel dashboard",
        sheets=sheets,
        confirmation_required=confirmation_required,
    )


def _style_sheet(worksheet) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center")
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row == 3:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for column_index, column_cells in enumerate(worksheet.columns, start=1):
        column_letter = get_column_letter(column_index)
        width = min(42, max(12, max(len(str(getattr(cell, "value", "") or "")) for cell in column_cells) + 2))
        worksheet.column_dimensions[column_letter].width = width


def _write_table(worksheet, table: WorkbookTablePlan) -> int:
    from openpyxl.utils.cell import coordinate_to_tuple

    start_row, start_col = coordinate_to_tuple(table.anchor)
    worksheet.cell(row=start_row - 1, column=start_col, value=table.title)
    for col_offset, header in enumerate(table.headers):
        worksheet.cell(row=start_row, column=start_col + col_offset, value=header)
    for row_offset, row in enumerate(table.rows, start=1):
        for col_offset, value in enumerate(row):
            worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
    return start_row + len(table.rows)


def _write_chart(worksheet, chart: WorkbookChartPlan) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils.cell import coordinate_to_tuple

    start_row, start_col = coordinate_to_tuple(chart.anchor)
    worksheet.cell(row=start_row, column=start_col, value="label")
    worksheet.cell(row=start_row, column=start_col + 1, value=chart.series_name or "value")
    for index, (label, value) in enumerate(zip(chart.labels, chart.values), start=1):
        worksheet.cell(row=start_row + index, column=start_col, value=label)
        worksheet.cell(row=start_row + index, column=start_col + 1, value=value)
    excel_chart = BarChart()
    excel_chart.title = chart.title
    excel_chart.add_data(
        Reference(worksheet, min_col=start_col + 1, min_row=start_row, max_row=start_row + len(chart.values)),
        titles_from_data=True,
    )
    excel_chart.set_categories(
        Reference(worksheet, min_col=start_col, min_row=start_row + 1, max_row=start_row + len(chart.labels))
    )
    worksheet.add_chart(excel_chart, "D3")


def render_workbook(plan: WorkbookPlan, output_path: str | Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise AnalysisError("openpyxl is required to generate Excel artifacts.") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_plan in plan.sheets:
        worksheet = workbook.create_sheet(sheet_plan.name)
        worksheet.merge_cells("A1:D1")
        worksheet["A1"] = sheet_plan.title or sheet_plan.name
        for table in sheet_plan.tables:
            _write_table(worksheet, table)
        for chart in sheet_plan.charts:
            _write_chart(worksheet, chart)
        for index, note in enumerate(sheet_plan.notes, start=1):
            worksheet.cell(row=index + 20, column=1, value=note)
        _style_sheet(worksheet)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def generate_excel_artifact(workspace: DocumentWorkspace) -> WorkspaceArtifact:
    from services import workspace_service

    plan = build_workbook_plan(workspace)
    workspace_dir = _ARTIFACT_ROOT / workspace.id
    filename = _safe_filename(plan.title)
    output_path = workspace_dir / filename
    render_workbook(plan, output_path)

    now = _utc_now()
    artifact = WorkspaceArtifact(
        id=f"artifact-{uuid4()}",
        workspace_id=workspace.id,
        kind="excel",
        filename=filename,
        content_type=EXCEL_CONTENT_TYPE,
        storage_path=str(output_path),
        plan=plan,
        warnings=list(plan.warnings),
        created_at=now,
        updated_at=now,
    )
    workspace.artifacts = [item for item in workspace.artifacts if item.kind != "excel"]
    workspace.artifacts.append(artifact)
    workspace_service.save_workspace(workspace)
    return artifact


def get_artifact(workspace_id: str, artifact_id: str) -> WorkspaceArtifact:
    from services import workspace_service

    workspace = workspace_service.get_workspace(workspace_id)
    artifact = next((item for item in workspace.artifacts if item.id == artifact_id), None)
    if artifact is None:
        raise AnalysisError("Workspace artifact was not found.")
    return artifact


def open_excel_artifact(workspace_id: str, artifact_id: str) -> WorkspaceArtifact:
    from services import workspace_service

    workspace = workspace_service.get_workspace(workspace_id)
    artifact = next((item for item in workspace.artifacts if item.id == artifact_id), None)
    if artifact is None:
        raise AnalysisError("Workspace artifact was not found.")
    if artifact.kind != "excel" or not artifact.storage_path:
        raise AnalysisError("Only generated Excel artifacts can be opened.")
    if os.name == "nt":
        os.startfile(artifact.storage_path)  # type: ignore[attr-defined]
        artifact.sync_state.status = "opened"
        artifact.sync_state.last_opened_at = _utc_now()
        artifact.updated_at = _utc_now()
        workspace_service.save_workspace(workspace)
        return artifact
    artifact.sync_state.status = "error"
    artifact.sync_state.error_message = "Excel auto-open is only available on Windows desktop."
    artifact.updated_at = _utc_now()
    workspace_service.save_workspace(workspace)
    return artifact


def _snapshot_workbook(path: str) -> dict:
    try:
        import openpyxl
    except ImportError as exc:
        raise AnalysisError("openpyxl is required to sync Excel artifacts.") from exc

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        sheets = {}
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows(max_row=30, max_col=12, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    rows.append(values)
            sheets[worksheet.title] = rows
        return {"source": "user_edit", "sheets": sheets}
    finally:
        workbook.close()


def sync_excel_artifact(workspace_id: str, artifact_id: str) -> WorkspaceArtifact:
    from services import workspace_service

    workspace = workspace_service.get_workspace(workspace_id)
    artifact = next((item for item in workspace.artifacts if item.id == artifact_id), None)
    if artifact is None:
        raise AnalysisError("Workspace artifact was not found.")
    if artifact.kind != "excel" or not artifact.storage_path:
        raise AnalysisError("Only generated Excel artifacts can be synced.")
    artifact.sync_state = WorkbookSyncState(
        status="synced",
        last_opened_at=artifact.sync_state.last_opened_at,
        last_synced_at=_utc_now(),
        snapshot=_snapshot_workbook(artifact.storage_path),
        warnings=list(artifact.sync_state.warnings),
    )
    artifact.updated_at = _utc_now()
    workspace_service.save_workspace(workspace)
    return artifact
