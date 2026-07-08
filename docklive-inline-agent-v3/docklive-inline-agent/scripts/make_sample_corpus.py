"""합성 샘플 양식 생성: 실제 양식 수집 전 채점기 자체를 검증하기 위한 표본.

- sample_apply.hwpx : 플레이스홀더 3개 + 표 구조를 가진 최소 유효 HWPX
- sample_budget.xlsx: 병합셀 + 수식 + 플레이스홀더 + 글자수 제한이 있는 예산표
"""
from __future__ import annotations

import json
import sys
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FORMS = ROOT / "corpus" / "forms"

HP = "http://www.hancom.co.kr/hwpml/2011/paragraph"

SECTION0 = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<hp:sec xmlns:hp="{HP}">
  <hp:p paraPrIDRef="0" styleIDRef="0"><hp:run charPrIDRef="0"><hp:t>사업 신청서</hp:t></hp:run></hp:p>
  <hp:p paraPrIDRef="1" styleIDRef="0"><hp:run charPrIDRef="1"><hp:t>기업명: {{{{회사명}}}}</hp:t></hp:run></hp:p>
  <hp:p paraPrIDRef="1" styleIDRef="0"><hp:run charPrIDRef="1"><hp:t>대표자: {{{{대표자}}}}</hp:t></hp:run></hp:p>
  <hp:p paraPrIDRef="1" styleIDRef="0"><hp:run charPrIDRef="1"><hp:t>사업 개요: {{{{사업개요}}}}</hp:t></hp:run></hp:p>
</hp:sec>'''

HEADER = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<hh:head xmlns:hh="http://www.hancom.co.kr/hwpml/2011/head" version="1.0">
  <hh:refList/>
</hh:head>'''

MANIFEST_XML = '''<?xml version="1.0" encoding="UTF-8"?>
<odf:manifest xmlns:odf="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0">
  <odf:file-entry odf:full-path="Contents/header.xml" odf:media-type="text/xml"/>
  <odf:file-entry odf:full-path="Contents/section0.xml" odf:media-type="text/xml"/>
</odf:manifest>'''


def make_hwpx(path: Path) -> None:
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(zipfile.ZipInfo("mimetype"), "application/hwp+zip",
                    compress_type=zipfile.ZIP_STORED)
        zf.writestr("META-INF/manifest.xml", MANIFEST_XML)
        zf.writestr("Contents/header.xml", HEADER)
        zf.writestr("Contents/section0.xml", SECTION0)


def make_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사업비"
    ws.merge_cells("A1:D1")
    ws["A1"] = "사업비 집행 계획"
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "비목", "산출근거", "금액", "비고"
    ws["A3"], ws["B3"], ws["C3"] = "인건비", "{{인건비산출}}", 12000000
    ws["A4"], ws["B4"], ws["C4"] = "재료비", "{{재료비산출}}", 3000000
    ws["C5"] = "=SUM(C3:C4)"
    ws["A5"] = "합계"
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18
    wb.save(path)


def main() -> int:
    FORMS.mkdir(parents=True, exist_ok=True)
    make_hwpx(FORMS / "sample_apply.hwpx")
    make_xlsx(FORMS / "sample_budget.xlsx")
    manifest = {
        "forms": [
            {"file": "sample_apply.hwpx", "source": "합성 표본", "program": "채점기 자가검증"},
            {"file": "sample_budget.xlsx", "source": "합성 표본", "program": "채점기 자가검증",
             "slots": [
                 {"id": "인건비산출", "kind": "cell", "ref": "사업비!B3", "max_chars": 50},
                 {"id": "재료비산출", "kind": "cell", "ref": "사업비!B4", "max_chars": 50},
             ]},
        ]
    }
    (ROOT / "corpus" / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"샘플 코퍼스 생성 완료: {FORMS}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
