"""무결성 채점기 자체 검증: 정상 통과 + '깨진 파일을 반드시 잡아내는가' 음성 테스트."""
import shutil
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from integrity import extract, fill, grader  # noqa: E402
from integrity.checks import hwpx_checks as hc  # noqa: E402
from integrity.checks import xlsx_checks as xc  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
FORMS = ROOT / "corpus" / "forms"
HWPX = FORMS / "sample_apply.hwpx"
XLSX = FORMS / "sample_budget.xlsx"


def test_roundtrip_passes_on_clean_corpus():
    summary = grader.grade_all()
    assert summary["total"] >= 2
    assert summary["pass_rate"] == 100.0, summary


def test_slot_extraction_finds_placeholders():
    slots = extract.extract_slots(HWPX)
    assert slots["ok"]
    ids = {s["id"] for s in slots["slots"]}
    assert {"회사명", "대표자", "사업개요"} <= ids


def test_grader_catches_bad_mimetype(tmp_path):
    """mimetype이 압축되거나 첫 엔트리가 아니면 반드시 FAIL이어야 한다."""
    bad = tmp_path / "bad.hwpx"
    with zipfile.ZipFile(HWPX) as zin, zipfile.ZipFile(bad, "w", zipfile.ZIP_DEFLATED) as zout:
        names = zin.namelist()
        for name in names:  # mimetype을 뒤로 + 압축해서 기록 (규약 위반 재현)
            if name != "mimetype":
                zout.writestr(name, zin.read(name))
        zout.writestr("mimetype", zin.read("mimetype"))
    ok, detail = hc.check_zip_valid(bad)
    assert ok is False and "mimetype" in detail


def test_grader_catches_structure_change(tmp_path):
    """채움이 태그 구조를 바꾸면(문단 삭제 등) C3가 잡아야 한다."""
    broken = tmp_path / "broken.hwpx"
    with zipfile.ZipFile(HWPX) as zin, zipfile.ZipFile(broken, "w") as zout:
        zout.writestr(zipfile.ZipInfo("mimetype"), zin.read("mimetype"),
                      compress_type=zipfile.ZIP_STORED)
        for name in zin.namelist():
            if name == "mimetype":
                continue
            data = zin.read(name)
            if name == "Contents/section0.xml":
                text = data.decode("utf-8")
                # 문단 하나 통째로 삭제 → 구조 변형
                start = text.index("<hp:p", text.index("대표자") - 200)
                end = text.index("</hp:p>", start) + len("</hp:p>")
                data = (text[:start] + text[end:]).encode("utf-8")
            zout.writestr(name, data)
    ok, detail = hc.check_structure_preserved(HWPX, broken)
    assert ok is False and "section0" in detail


def test_grader_catches_leftover_placeholder(tmp_path):
    """일부 슬롯만 채우면 C5(미채움)가 잡아야 한다."""
    dst = tmp_path / "partial.hwpx"
    slots = extract.extract_slots(HWPX)["slots"]
    fill.fill_file(HWPX, dst, slots[:1])  # 1개만 채움
    ok, detail = hc.check_no_placeholder_left(dst)
    assert ok is False and "{{" in detail


def test_grader_catches_broken_merge(tmp_path):
    """병합 셀이 사라지면 C3가 잡아야 한다 — 양식 깨짐의 최다 원인."""
    import openpyxl
    broken = tmp_path / "broken.xlsx"
    shutil.copy2(XLSX, broken)
    wb = openpyxl.load_workbook(broken)
    wb["사업비"].unmerge_cells("A1:D1")
    wb.save(broken)
    ok, detail = xc.check_merged_cells_preserved(XLSX, broken)
    assert ok is False and "병합" in detail


def test_grader_catches_clobbered_formula(tmp_path):
    """채움 대상이 아닌 수식 셀이 값으로 덮어써지면 C4가 잡아야 한다."""
    import openpyxl
    broken = tmp_path / "broken2.xlsx"
    shutil.copy2(XLSX, broken)
    wb = openpyxl.load_workbook(broken)
    wb["사업비"]["C5"] = 15000000  # =SUM(...) 파괴
    wb.save(broken)
    ok, detail = xc.check_formulas_preserved(XLSX, broken, fill_refs=set())
    assert ok is False and "C5" in detail


def test_char_limit_enforced(tmp_path):
    """max_chars를 넘겨 채우면 C5(글자수)가 잡아야 한다."""
    import openpyxl
    over = tmp_path / "over.xlsx"
    shutil.copy2(XLSX, over)
    wb = openpyxl.load_workbook(over)
    wb["사업비"]["B3"] = "가" * 100
    wb.save(over)
    slots = [{"id": "인건비산출", "kind": "cell", "ref": "사업비!B3", "max_chars": 50}]
    ok, detail = xc.check_char_limits(over, slots)
    assert ok is False and "초과" in detail
