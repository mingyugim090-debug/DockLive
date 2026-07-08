"""XLSX 라운드트립 무결성 검사."""
from __future__ import annotations

from pathlib import Path

import openpyxl


def check_loads(path: Path) -> tuple[bool, str]:
    """C1: openpyxl로 열리는가 (파일 무손상)."""
    try:
        openpyxl.load_workbook(path)
        return True, "정상 로드"
    except Exception as e:
        return False, f"로드 실패: {e}"


def check_sheets_preserved(original: Path, filled: Path) -> tuple[bool, str]:
    """C2: 시트 이름/순서 보존."""
    o = openpyxl.load_workbook(original).sheetnames
    f = openpyxl.load_workbook(filled).sheetnames
    return (o == f, "시트 보존" if o == f else f"시트 변경: {o} → {f}")


def check_merged_cells_preserved(original: Path, filled: Path) -> tuple[bool, str]:
    """C3: 병합 셀 범위 보존 — 정부 양식 깨짐의 최다 원인."""
    wo, wf = openpyxl.load_workbook(original), openpyxl.load_workbook(filled)
    for name in wo.sheetnames:
        mo = {str(r) for r in wo[name].merged_cells.ranges}
        mf = {str(r) for r in wf[name].merged_cells.ranges}
        if mo != mf:
            return False, f"'{name}' 병합 변경: 사라짐 {sorted(mo - mf)[:3]}, 생김 {sorted(mf - mo)[:3]}"
    return True, "병합 셀 보존"


def check_formulas_preserved(original: Path, filled: Path,
                             fill_refs: set[str]) -> tuple[bool, str]:
    """C4: 채움 대상이 아닌 수식 셀이 덮어써지지 않았는가."""
    wo = openpyxl.load_workbook(original, data_only=False)
    wf = openpyxl.load_workbook(filled, data_only=False)
    broken = []
    for name in wo.sheetnames:
        for row in wo[name].iter_rows():
            for cell in row:
                ref = f"{name}!{cell.coordinate}"
                if (isinstance(cell.value, str) and cell.value.startswith("=")
                        and ref not in fill_refs):
                    if wf[name][cell.coordinate].value != cell.value:
                        broken.append(ref)
    return (not broken, "수식 보존" if not broken else f"수식 파손: {broken[:5]}")


def check_char_limits(filled: Path, slots: list[dict]) -> tuple[bool, str]:
    """C5: max_chars가 지정된 슬롯의 글자 수 제한 준수."""
    wf = openpyxl.load_workbook(filled)
    over = []
    for s in slots:
        limit = s.get("max_chars")
        if not limit or s["kind"] != "cell":
            continue
        name, coord = s["ref"].split("!")
        val = wf[name][coord].value or ""
        if len(str(val)) > limit:
            over.append(f"{s['ref']}({len(str(val))}>{limit})")
    return (not over, "글자수 제한 준수" if not over else f"초과: {over}")
