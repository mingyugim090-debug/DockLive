import sys
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from tools import integrity_tools  # noqa: E402


def _write_xlsx(path: Path, *, formula_value: str = "=SUM(A1:B1)", placeholder: bool = False) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 1
    ws["B1"] = 2
    ws["C1"] = formula_value
    ws.merge_cells("A3:B3")
    ws["A3"] = "{{applicant}}" if placeholder else "Applicant"
    wb.save(path)


def _write_hwpx(path: Path, *, malformed_xml: bool = False, bad_mimetype: bool = False) -> None:
    content = b"<root><p>Done</p></root>"
    if malformed_xml:
        content = b"<root><p>Broken</root>"
    mimetype = b"application/octet-stream" if bad_mimetype else b"application/hwp+zip"

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(zipfile.ZipInfo("mimetype"), mimetype, compress_type=zipfile.ZIP_STORED)
        zf.writestr("Contents/section0.xml", content)


def _check(summary: dict, check_id: str) -> dict:
    return next(check for check in summary["checks"] if check["id"] == check_id)


def test_validate_xlsx_reports_leftover_placeholders(tmp_path):
    path = tmp_path / "filled.xlsx"
    _write_xlsx(path, placeholder=True)

    out = integrity_tools.validate_document(str(path))

    assert out["ok"] is True
    assert out["data"]["validation_passed"] is False
    assert _check(out["data"], "no_placeholders")["passed"] is False


def test_validate_xlsx_detects_formula_changed_outside_authored_ranges(tmp_path):
    original = tmp_path / "original.xlsx"
    filled = tmp_path / "filled.xlsx"
    _write_xlsx(original)
    _write_xlsx(filled, formula_value=99)

    out = integrity_tools.validate_document(
        str(filled),
        original_path=str(original),
        authored_ranges=["Data!A1:B1"],
    )

    assert out["ok"] is True
    assert out["data"]["validation_passed"] is False
    assert _check(out["data"], "formulas_preserved")["passed"] is False
    assert "Data!C1" in _check(out["data"], "formulas_preserved")["detail"]


def test_validate_hwpx_checks_mimetype_and_xml(tmp_path):
    good = tmp_path / "good.hwpx"
    bad_mimetype = tmp_path / "bad-mimetype.hwpx"
    malformed = tmp_path / "malformed.hwpx"
    _write_hwpx(good)
    _write_hwpx(bad_mimetype, bad_mimetype=True)
    _write_hwpx(malformed, malformed_xml=True)

    good_out = integrity_tools.validate_document(str(good))
    bad_mimetype_out = integrity_tools.validate_document(str(bad_mimetype))
    malformed_out = integrity_tools.validate_document(str(malformed))

    assert good_out["ok"] is True
    assert good_out["data"]["validation_passed"] is True
    assert _check(bad_mimetype_out["data"], "hwpx_mimetype")["passed"] is False
    assert _check(malformed_out["data"], "xml_well_formed")["passed"] is False
