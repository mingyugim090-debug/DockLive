import json
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from integrity import extract, fill, grader  # noqa: E402
from integrity.checks import hwpx_checks, xlsx_checks  # noqa: E402


FORMS = ROOT / "corpus" / "forms"
SAMPLE_HWPX = FORMS / "sample_apply.hwpx"
SAMPLE_XLSX = FORMS / "sample_budget.xlsx"


def test_integrity_corpus_manifest_is_registered():
    manifest = json.loads((ROOT / "corpus" / "manifest.json").read_text(encoding="utf-8"))

    files = {entry["file"] for entry in manifest["forms"]}

    assert {"sample_apply.hwpx", "sample_budget.xlsx"} <= files


def test_roundtrip_passes_on_registered_corpus():
    summary = grader.grade_all()

    assert summary["total"] >= 2
    assert summary["pass_rate"] == 100.0, summary


def test_grade_file_cli_writes_failure_report():
    result = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "grade.py"), "--file", "sample_budget.xlsx"],
        cwd=ROOT,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    assert (ROOT / "workspace" / "reports" / "integrity_latest.md").exists()
    assert (ROOT / "workspace" / "reports" / "integrity_latest.json").exists()


def test_write_report_appends_history_jsonl(tmp_path, monkeypatch):
    monkeypatch.setattr(grader, "REPORTS", tmp_path)
    first = {
        "timestamp": "2026-07-08T10:00:00",
        "total": 1,
        "passed": 1,
        "pass_rate": 100.0,
        "results": [],
    }
    second = dict(first, timestamp="2026-07-08T10:01:00")

    grader.write_report(first)
    grader.write_report(second)

    lines = (tmp_path / "history.jsonl").read_text(encoding="utf-8").splitlines()
    assert len(lines) == 2
    assert json.loads(lines[-1])["timestamp"] == "2026-07-08T10:01:00"


def test_slot_extraction_finds_hwpx_placeholders():
    result = extract.extract_slots(SAMPLE_HWPX)

    assert result["ok"], result
    ids = {slot["id"] for slot in result["slots"]}
    assert {"회사명", "대표자", "사업개요"} <= ids


def test_grader_catches_bad_mimetype(tmp_path):
    bad = tmp_path / "bad.hwpx"
    with zipfile.ZipFile(SAMPLE_HWPX) as zin, zipfile.ZipFile(bad, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            if name != "mimetype":
                zout.writestr(name, zin.read(name))
        zout.writestr("mimetype", zin.read("mimetype"))

    ok, detail = hwpx_checks.check_zip_valid(bad)

    assert ok is False
    assert "mimetype" in detail


def test_grader_catches_leftover_placeholder(tmp_path):
    partial = tmp_path / "partial.hwpx"
    slots = extract.extract_slots(SAMPLE_HWPX)["slots"]
    fill.fill_file(SAMPLE_HWPX, partial, slots[:1])

    ok, detail = hwpx_checks.check_no_placeholder_left(partial)

    assert ok is False
    assert "{{" in detail


def test_grader_catches_broken_merge(tmp_path):
    broken = tmp_path / "broken.xlsx"
    shutil.copy2(SAMPLE_XLSX, broken)
    wb = openpyxl.load_workbook(broken)
    ws = wb["사업비"]
    ws.unmerge_cells("A1:D1")
    wb.save(broken)

    ok, detail = xlsx_checks.check_merged_cells_preserved(SAMPLE_XLSX, broken)

    assert ok is False
    assert "merged" in detail.lower()


def test_grader_catches_clobbered_formula(tmp_path):
    broken = tmp_path / "formula.xlsx"
    shutil.copy2(SAMPLE_XLSX, broken)
    wb = openpyxl.load_workbook(broken)
    wb["사업비"]["C5"] = 15000000
    wb.save(broken)

    ok, detail = xlsx_checks.check_formulas_preserved(SAMPLE_XLSX, broken, fill_refs=set())

    assert ok is False
    assert "C5" in detail


def test_grader_enforces_manifest_char_limit(tmp_path):
    over = tmp_path / "over.xlsx"
    shutil.copy2(SAMPLE_XLSX, over)
    wb = openpyxl.load_workbook(over)
    wb["사업비"]["B3"] = "A" * 100
    wb.save(over)

    slots = [{"id": "인건비산출", "kind": "cell", "ref": "사업비!B3", "max_chars": 50}]
    ok, detail = xlsx_checks.check_char_limits(over, slots)

    assert ok is False
    assert "exceeds" in detail.lower()
