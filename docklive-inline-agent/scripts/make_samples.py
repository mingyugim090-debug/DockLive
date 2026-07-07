"""데모용 견적서 양식 생성 (Phase 3).

samples/견적서양식.xlsx 를 만든다. samples/originals/ 는 건드리지 않는다.
사용: python scripts/make_samples.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

SAMPLES_DIR = Path(__file__).resolve().parents[1] / "samples"

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="DDE7E2")


def make_quotation_form(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    ws["B2"] = "견 적 서"
    ws["B2"].font = Font(size=18, bold=True)
    ws["B4"] = "수신:"
    ws["B5"] = "견적일자:"
    ws["B6"] = "유효기간: 견적일로부터 30일"

    headers = ["번호", "품목", "수량", "단가", "금액"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row in range(9, 14):  # 품목 5행 (빈 양식)
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = BORDER

    ws["E14"] = "합계"
    ws["E14"].font = Font(bold=True)
    ws["E14"].border = BORDER
    ws["F14"].border = BORDER

    for col, width in zip("BCDEF", (8, 28, 8, 14, 16)):
        ws.column_dimensions[col].width = width

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"생성: {path}")


if __name__ == "__main__":
    make_quotation_form(SAMPLES_DIR / "견적서양식.xlsx")
