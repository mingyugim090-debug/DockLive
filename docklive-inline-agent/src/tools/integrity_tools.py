"""Read-only integrity checks for completed local document outputs."""
from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries


PLACEHOLDER_RE = re.compile(r"\{\{[^{}]+\}\}")


def _err(msg: str) -> dict:
    return {"ok": False, "error": msg}


def _ok(data) -> dict:
    return {"ok": True, "data": data}


def _check(check_id: str, label: str, passed: bool, detail: str = "") -> dict:
    return {"id": check_id, "label": label, "passed": bool(passed), "detail": detail}


def _summary(path: Path, file_format: str, checks: list[dict], warnings: list[str] | None = None) -> dict:
    return {
        "path": str(path),
        "format": file_format,
        "validation_passed": bool(checks) and all(check["passed"] for check in checks),
        "checks": checks,
        "warnings": warnings or [],
    }


def validate_document(
    path: str,
    original_path: str = "",
    authored_ranges: list[str] | None = None,
) -> dict:
    """Validate a saved document without mutating it."""
    target = Path(path)
    if not target.exists():
        return _err(f"file does not exist: {path}")

    suffix = target.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _ok(_validate_xlsx(target, Path(original_path) if original_path else None, authored_ranges or []))
    if suffix == ".hwpx":
        return _ok(_validate_hwpx(target))
    if suffix == ".hwp":
        return _ok(
            _summary(
                target,
                "hwp",
                [_check("unsupported_hwp_binary", "HWP binary validation", False, "HWP binary validation is not available locally")],
                ["Use backend HWP-to-HWPX conversion/export for full validation."],
            )
        )
    return _ok(
        _summary(
            target,
            suffix.lstrip(".") or "unknown",
            [_check("unsupported_format", "Supported document format", False, f"unsupported format: {suffix}")],
        )
    )


def _validate_xlsx(path: Path, original_path: Path | None, authored_ranges: list[str]) -> dict:
    checks: list[dict] = []
    warnings: list[str] = []
    try:
        workbook = openpyxl.load_workbook(path, data_only=False)
        checks.append(_check("xlsx_loads", "Workbook loads", True, "openpyxl loaded workbook"))
    except Exception as exc:
        return _summary(path, "xlsx", [_check("xlsx_loads", "Workbook loads", False, str(exc))])

    checks.append(_check("no_placeholders", "No leftover placeholders", *_no_xlsx_placeholders(workbook)))

    original = None
    if original_path and original_path.exists():
        try:
            original = openpyxl.load_workbook(original_path, data_only=False)
        except Exception as exc:
            warnings.append(f"Could not load original workbook for comparison: {exc}")
    elif original_path:
        warnings.append(f"Original workbook does not exist: {original_path}")

    if original is not None:
        checks.append(_check("sheets_preserved", "Sheets preserved", *_sheets_preserved(original, workbook)))
        checks.append(_check("merged_cells_preserved", "Merged cells preserved", *_merged_cells_preserved(original, workbook)))
        authored_refs = _expand_authored_ranges(authored_ranges)
        checks.append(_check("formulas_preserved", "Formulas preserved", *_formulas_preserved(original, workbook, authored_refs)))

    return _summary(path, "xlsx", checks, warnings)


def _no_xlsx_placeholders(workbook) -> tuple[bool, str]:
    leftovers: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and PLACEHOLDER_RE.search(cell.value):
                    leftovers.append(f"{sheet.title}!{cell.coordinate}")
    if leftovers:
        return False, ", ".join(leftovers[:10])
    return True, "no placeholders found"


def _sheets_preserved(original, filled) -> tuple[bool, str]:
    original_names = original.sheetnames
    filled_names = filled.sheetnames
    if original_names == filled_names:
        return True, "sheet names and order preserved"
    return False, f"{original_names} -> {filled_names}"


def _merged_cells_preserved(original, filled) -> tuple[bool, str]:
    changed: list[str] = []
    for sheet_name in original.sheetnames:
        if sheet_name not in filled.sheetnames:
            changed.append(f"{sheet_name}: missing sheet")
            continue
        original_ranges = {str(item) for item in original[sheet_name].merged_cells.ranges}
        filled_ranges = {str(item) for item in filled[sheet_name].merged_cells.ranges}
        if original_ranges != filled_ranges:
            changed.append(f"{sheet_name}: {sorted(original_ranges)} -> {sorted(filled_ranges)}")
    if changed:
        return False, "; ".join(changed[:5])
    return True, "merged cells preserved"


def _formulas_preserved(original, filled, authored_refs: set[str]) -> tuple[bool, str]:
    changed: list[str] = []
    for sheet_name in original.sheetnames:
        if sheet_name not in filled.sheetnames:
            continue
        for row in original[sheet_name].iter_rows():
            for cell in row:
                ref = f"{sheet_name}!{cell.coordinate}"
                if ref in authored_refs:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if filled[sheet_name][cell.coordinate].value != cell.value:
                        changed.append(ref)
    if changed:
        return False, ", ".join(changed[:10])
    return True, "formulas preserved"


def _expand_authored_ranges(ranges: list[str]) -> set[str]:
    refs: set[str] = set()
    for item in ranges:
        if "!" not in item:
            continue
        sheet_name, coord = item.split("!", 1)
        sheet_name = sheet_name.strip("'")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coord)
        except ValueError:
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                refs.add(f"{sheet_name}!{get_column_letter(col)}{row}")
    return refs


def _validate_hwpx(path: Path) -> dict:
    checks: list[dict] = []
    try:
        with zipfile.ZipFile(path) as zf:
            infos = zf.infolist()
            mimetype_ok = bool(infos) and infos[0].filename == "mimetype" and infos[0].compress_type == zipfile.ZIP_STORED
            if mimetype_ok:
                mimetype_ok = zf.read("mimetype").decode("utf-8", errors="ignore").strip() == "application/hwp+zip"
            checks.append(_check("hwpx_mimetype", "HWPX mimetype", mimetype_ok, "mimetype is first and stored" if mimetype_ok else "invalid mimetype entry"))
            checks.append(_check("xml_well_formed", "XML well-formed", *_hwpx_xml_well_formed(zf)))
            checks.append(_check("no_placeholders", "No leftover placeholders", *_no_hwpx_placeholders(zf)))
    except zipfile.BadZipFile:
        checks.append(_check("hwpx_zip", "HWPX zip package", False, "not a valid zip package"))
    except Exception as exc:
        checks.append(_check("hwpx_package", "HWPX package", False, str(exc)))
    return _summary(path, "hwpx", checks)


def _hwpx_xml_well_formed(zf: zipfile.ZipFile) -> tuple[bool, str]:
    bad: list[str] = []
    for name in zf.namelist():
        if name.endswith(".xml") or name.endswith(".hpf"):
            try:
                ET.fromstring(zf.read(name))
            except ET.ParseError as exc:
                bad.append(f"{name}: {exc}")
    if bad:
        return False, "; ".join(bad[:5])
    return True, "all XML entries parsed"


def _no_hwpx_placeholders(zf: zipfile.ZipFile) -> tuple[bool, str]:
    leftovers: list[str] = []
    for name in zf.namelist():
        if name.startswith("Contents/section") and name.endswith(".xml"):
            text = zf.read(name).decode("utf-8", errors="ignore")
            leftovers.extend(PLACEHOLDER_RE.findall(text))
    if leftovers:
        return False, ", ".join(sorted(set(leftovers))[:10])
    return True, "no placeholders found"
