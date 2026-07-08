from __future__ import annotations

from pathlib import Path

import openpyxl


def check_loads(path: Path) -> tuple[bool, str]:
    try:
        openpyxl.load_workbook(path)
    except Exception as exc:
        return False, f"workbook load failed: {exc}"
    return True, "workbook loads"


def check_sheets_preserved(original: Path, filled: Path) -> tuple[bool, str]:
    original_names = openpyxl.load_workbook(original).sheetnames
    filled_names = openpyxl.load_workbook(filled).sheetnames
    if original_names != filled_names:
        return False, f"sheet names changed: {original_names} -> {filled_names}"
    return True, "sheets are preserved"


def check_merged_cells_preserved(original: Path, filled: Path) -> tuple[bool, str]:
    original_wb = openpyxl.load_workbook(original)
    filled_wb = openpyxl.load_workbook(filled)
    for sheet_name in original_wb.sheetnames:
        original_ranges = {str(item) for item in original_wb[sheet_name].merged_cells.ranges}
        filled_ranges = {str(item) for item in filled_wb[sheet_name].merged_cells.ranges}
        if original_ranges != filled_ranges:
            return False, (
                f"merged cells changed in {sheet_name}: "
                f"removed={sorted(original_ranges - filled_ranges)}, added={sorted(filled_ranges - original_ranges)}"
            )
    return True, "merged cells are preserved"


def check_formulas_preserved(original: Path, filled: Path, fill_refs: set[str]) -> tuple[bool, str]:
    original_wb = openpyxl.load_workbook(original, data_only=False)
    filled_wb = openpyxl.load_workbook(filled, data_only=False)
    broken: list[str] = []
    for sheet_name in original_wb.sheetnames:
        for row in original_wb[sheet_name].iter_rows():
            for cell in row:
                ref = f"{sheet_name}!{cell.coordinate}"
                if ref in fill_refs:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if filled_wb[sheet_name][cell.coordinate].value != cell.value:
                        broken.append(ref)
    if broken:
        return False, f"formulas changed: {broken[:5]}"
    return True, "formulas are preserved"


def check_char_limits(filled: Path, slots: list[dict]) -> tuple[bool, str]:
    workbook = openpyxl.load_workbook(filled)
    exceeded: list[str] = []
    for slot in slots:
        limit = slot.get("max_chars")
        if not isinstance(limit, int) or limit <= 0 or slot.get("kind") != "cell":
            continue
        sheet_name, coord = str(slot["ref"]).split("!", 1)
        value = workbook[sheet_name][coord].value or ""
        length = len(str(value))
        if length > limit:
            exceeded.append(f"{slot['ref']} exceeds {limit} chars ({length})")
    if exceeded:
        return False, "; ".join(exceeded[:5])
    return True, "character limits are respected"
